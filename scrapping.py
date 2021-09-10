import openpyxl as xl
import os

currentPath = os.path.dirname(__file__)
reportPath = os.path.join(currentPath, "실습해보자임마")

print(currentPath)
print(reportPath)

reports = []

for file in os.listdir(reportPath):
    if file.endswith("xlsx"): # and ("찐" in file):
        print(file)
        filePath = os.path.join(reportPath, file)

        wb = xl.load_workbook(filePath)
        sheet = wb.active

        row = 1
        col = 1

        for i in range(10):
            name = sheet.cell(row= row, column = col).value
            if name is None:
                break
            first = sheet.cell(row= row + 1, column = col + 1).value
            second = sheet.cell(row= row + 1, column = col + 2).value
            third = sheet.cell(row= row + 1, column = col + 3).value
            fourth = sheet.cell(row= row + 1, column = col + 4).value
            fifth = sheet.cell(row= row + 1, column = col + 5).value

            reports.append({"name":name, "first":first, "second":second, "third":third, "fourth":fourth, "fifth":fifth})
            row += 1

print(reports)